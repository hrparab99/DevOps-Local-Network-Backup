import time
from openpyxl import load_workbook
from ping3 import ping
import os
import subprocess

def is_ip_responsive(ip):
    """Ping the IP address and check if it is responsive"""
    response = ping(ip, timeout=2)
    if response is None:
        return False
    return True

def check_hostnames_and_run_scripts(excel_file):
    wb = load_workbook(excel_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True):
        hostname, source_path, destination_path, password = row
        # Ensure source and destination paths exist
        # if not os.path.exists(source_path):
        #     print(f"Source path {source_path} does not exist. Skipping.")
        #     continue
        
        print(f"Pinging {hostname}...")
        # # Check if hostname is responsive
        if is_ip_responsive(hostname):
            print(f"{hostname} is responsive. Initializing Backup:")
            if not os.path.exists(source_path):
                print(f"{source_path} path not exist. Skipping...")
            else:
                print(f"Source path: {source_path}")
                print(f"Destination path: {destination_path}")

            rsync_command = f"sshpass -p '{password}' rsync -avu -e 'ssh -o StrictHostKeyChecking=no' {source_path}/ {destination_path}/"
            try:
                # Execute rsync command
                result = subprocess.run(
                    rsync_command,
                    shell=True,
                    check=True,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True,
                )
                print(result.stdout)
                print("Backup completed successfully.")
            except subprocess.CalledProcessError as e:
                print(f"Error during rsync execution: {e.stderr}")
        else:
            print(f"{hostname} is not responsive. Skipping Backup.")

def run_periodically(excel_file, interval):
    while True:
        print("Checking Excel sheet and starting backup...")
        check_hostnames_and_run_scripts(excel_file)
        print(f"Waiting for {interval} seconds before the next run...")
        time.sleep(interval)  # Wait for the specified interval (e.g., 600 seconds or 10 minutes)

if __name__ == "__main__":
    excel_file = os.getenv("EXCEL_FILE", "hostnames_and_scripts.xlsx")  # Use env var or default
    interval = int(os.getenv("INTERVAL", "600"))  # Interval in seconds
    run_periodically(excel_file, interval)  # Run every 10 minutes